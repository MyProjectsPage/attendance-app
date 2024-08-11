#####################################################################################
# VINAVIL ATTENDANCE APPLICATION
# SHADY555@GMAIL.COM
# AUGUST2024
#####################################################################################


#####################################################################################
# BACKEND CALCULATION
#####################################################################################


#import os
#import subprocess
import pandas as pd
#import win32com.client
pd.set_option('display.max_rows', None, 'display.max_columns', None, 'display.max_colwidth', None)
pd.set_option('display.precision', 2)
pd.options.display.width = 0



def create_xl_file(df_or_dfs, output_file='Output.xlsx', open_file=False, sheet_names=None):
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl import load_workbook
    from openpyxl.styles import Font
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.worksheet.table import Table, TableStyleInfo
    from io import BytesIO
    import os
    import win32com.client

    # If df_or_dfs is not a list, convert it to a list with a single item
    if not isinstance(df_or_dfs, list):
        df_or_dfs = [df_or_dfs]

    # If sheet_names is provided, ensure it's the same length as df_or_dfs
    if sheet_names and len(sheet_names) != len(df_or_dfs):
        raise ValueError("Length of sheet_names must match the number of DataFrames")

    # Create a new workbook or load the existing one
    wb = Workbook()
    wb.remove(wb.active)  # Remove the default sheet created with the workbook

    for i, df in enumerate(df_or_dfs):
        # Determine the sheet name
        sheet_name = sheet_names[i] if sheet_names else f'Sheet{i+1}'
        
        # Create a new sheet
        ws = wb.create_sheet(title=sheet_name)

        # Write the DataFrame to the sheet
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True)):
            ws.append(row)

        # Create a table and add it to the worksheet
        tab = Table(displayName=f"Table{i+1}", ref=ws.dimensions)
        style = TableStyleInfo(name="TableStyleLight9", showFirstColumn=False, showLastColumn=False, showRowStripes=False, showColumnStripes=True)
        tab.tableStyleInfo = style
        ws.add_table(tab)
        ws.freeze_panes = 'A2'

        # Set font to Calibri size 9 for all cells
        font = Font(name='Calibri', size=9)
        for row in ws.iter_rows():
            for cell in row:
                cell.font = font

        # Set header font to white
        white_font = Font(color="FFFFFF")
        for cell in ws[1]:  # First header row
            cell.font = Font(name='Calibri', size=9, color="FFFFFF")

        # Auto-fit column widths
        from openpyxl.utils.cell import get_column_letter
        for column_cells in ws.columns:
            new_column_length = max(len(str(cell.value)) for cell in column_cells)
            new_column_letter = (get_column_letter(column_cells[0].column))
            if new_column_length > 0:
                ws.column_dimensions[new_column_letter].width = new_column_length*1.23
            

    # Save the workbook as an Excel file
    wb.save(output_file)

    # Save the workbook in memory for Streamlit download button
    output_file_stream = BytesIO()
    wb.save(output_file_stream)
    output_file_stream.seek(0)

    # Open Excel application
    if open_file:
        script_dir = os.path.dirname(os.path.abspath(__file__))
        output_file_path = os.path.join(script_dir, output_file)
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = True  # Make Excel visible
        workbook = excel.Workbooks.Open(output_file_path)

    return output_file_stream





def delete_irrelivant_entries(df):

    # Initialize an empty list to store the indices to keep
    indices_to_keep = [0]  # Always keep the first row

    # Loop through the DataFrame starting from the second row
    for i in range(1, len(df)):
        # Compare the current row's 'state' with the previous row's 'state'
        if df.iloc[i]['state'] != df.iloc[i - 1]['state']: indices_to_keep.append(i)

    # Create a new DataFrame with the rows to keep
    df = df.iloc[indices_to_keep].reset_index(drop=True)
    return df

def merge_ins_outs(df):
    df_ins = df.iloc[::2].reset_index(drop=True)  # odd index
    df_outs  = df.iloc[1::2].reset_index(drop=True)  # even index
    if len(df_ins) != len(df_outs): raise
    
    df = df_ins[['id', 'name']]
    df.loc[:, ['state_in', 'date_time_in']] = df_ins[['state', 'date_time']].values
    df.loc[:, ['state_out', 'date_time_out']] = df_outs[['state', 'date_time']].values
    return df

def calc_time_spent(df):
    df = df.copy(deep=True) # To avoid copy warning
    df['time_spent'] = df['date_time_out'] - df['date_time_in']
    df['hours'] = df['time_spent'].dt.total_seconds() / 3600
    df['time_spent'] = df['time_spent'].astype(str)
    df['time_spent'] = df['time_spent'].str[-8:-3]
    df['hours'] = df['hours'].round(1)

    for item in ['date_time_in', 'date_time_out']:
        df[item] = df[item].dt.strftime('%d-%m-%Y %I:%M %p')  # AM PM Format
    df['name'] = df['name'].str.upper()
    df.insert(2, 'shift', range(1, len(df) + 1))
    df.insert(3, 'work day', df['date_time_in'])
    df['work day'] = pd.to_datetime(df['work day'], format='%d-%m-%Y %I:%M %p').dt.strftime('%d-%a').str.upper()
    df.columns = df.columns.str.upper()
    
    return df

    

def run_backend(df):

    #####################################################################################
    # BCKEND PROCESSING - MAIN
    #####################################################################################

    dfr = pd.DataFrame()
    #df = pd.read_excel('attendance.xlsx')
    df.columns = 'id name date_time state'.split()
    #df = df[df['name'] == 'Adnan Mohamed Abdul Hami']

    df['date_time'] = pd.to_datetime(df['date_time'], format='%m/%d/%Y %I:%M %p')  
    df = df.sort_values(by=df.columns.tolist())
    df = df[['id', 'name', 'state', 'date_time']]
    df_all = df.copy(deep=True)

    names = (df['name'].unique())

    for name in names:
        df = df_all[df_all['name'] == name]
        while df.iloc[0]['state']  != 'C/In': df = df.iloc[1:].reset_index(drop=True)
        while df.iloc[-1]['state'] != 'C/Out': df = df.iloc[:-1].reset_index(drop=True)
        df = delete_irrelivant_entries(df)
        df = merge_ins_outs(df)
        df = calc_time_spent(df)
        dfr = pd.concat([dfr, df], ignore_index=True)

    output_file_stream = create_xl_file([df_all, dfr], sheet_names= ['Input', 'Output'])
    print('\n'*3)
    print(dfr)

    # Abnormal hours (e.g. someone working 23 hours)
    df_abnormal = dfr[~dfr['HOURS'].between(2, 11)]

    return dfr, df_abnormal, output_file_stream










#####################################################################################
# FRONTEND STREAMLIT APP
#####################################################################################


import streamlit as st
import pandas as pd


# Set page configuration
st.set_page_config(page_title='Attendance Calculator', page_icon='📊', layout='wide', initial_sidebar_state='expanded')

# Use dark theme and set background image from URL
st.markdown(
    """
    <style>
    .reportview-container {
        background: none;
    }
    .main {
        background: url('https://wallpapercave.com/wp/wp4312403.jpg') no-repeat center center fixed;
        background: url('https://img.freepik.com/premium-photo/abstract-wave-element-design-blue-curve-light-lines-background-digital-frequency-track-equalizer-generative-ai_1423-11938.jpg?w=826') no-repeat center center fixed;
        background: url('https://wallpapercave.com/wp/wp4312424.jpg') no-repeat center center fixed;
        background-size: cover;
    }
    </style>
    """,
    unsafe_allow_html=True
)

# Custom CSS to create the navigation bar
st.markdown("""
    <style>
    .navbar {
        display: flex;
        justify-content: flex-end;
        background-color: transparent;  /* Transparent background */
        padding: 0.5rem 1rem;  /* Less padding to move it closer to the top */
        margin-top: -3rem;     /* Negative margin to push it further up */
        padding-right: 10rem;   /* Adjust padding-right to move the navbar to the left */
        position: fixed;       /* Fix the navbar at the top */
        width: 100%;           /* Make sure the navbar spans the width of the page */
        /* z-index: 1000;         /* Ensure the navbar is above other content */
    }
    .navbar a {
        color: navy blue;
        text-decoration: none;
        padding: 0.5rem 1rem;
        font-size: 1.2rem;
    }
    .navbar a:hover {
        color: #007BFF; /* Change to a color you like */
    }
    </style>
    <div class="navbar">
        <a href="/?nav=attendance-calculator">Attendance Calculator</a>
        <a href="/?nav=about">About</a>
    </div>
    """, unsafe_allow_html=True)

# Get the query parameters to determine which page to show
query_params = st.query_params
page = query_params.get("nav", "attendance-calculator")





############################################
# PAGES FOR ABOUT AND ATTENDANCE CALCULATOR
############################################


# Initialize the page with "Attendance Calculator" as the default
if "page" not in st.session_state:
    st.session_state.page = "Attendance Calculator"


# Define the content for each page
if page == "about":
    st.markdown("## About")
    st.markdown("Developer: Chadee Fouad - MyWorkDropBox@gmail.com  \nDevelopment Date: Aug 2024.")
    st.write("")
    text = 'The purpose of this application is to help payroll accountants at a factory to calculate the attendance for workers.'
    text = text + '  \nCurrently there are many issues which makes the process quite complicated and requires a lot of manual adjustments.'
    text = text + '  \nThe reason for this is that the current attendance scanner is very basic.'
    text = text + "  \nAs such anyone can 'check-in' or 'check-out' many times. For example an employee can check in at 9:00 AM then go get something from his car then check in again at 9:03 AM."
    text = text + "  \nOften there are people that are worried that the scanner did not scan correctly so they check-in or check-out many times when the scanner already scanned correctly."
    text = text + "  \nSometimes people forget to check-in or check-out so the in/outs do not correctly align."
    text = text + "  \nThis is where the application helps. It gives tries to align those entries as much as possible."
    text = text + "  \nIt also highlights shifts with an abnormal number of hours (possible check-in/out errors) which makes it much easier for the accountants."
    st.markdown(text)
    

elif page == "attendance-calculator":
    
    # Title
    st.title('Attendance Calculator')

    # File upload
    uploaded_file = st.file_uploader("Upload your file", type=["csv", "xlsx"])

    if uploaded_file is not None:
        # Read file into dataframe
        if uploaded_file.name.endswith('.csv'):
            df1 = pd.read_csv(uploaded_file)
        elif uploaded_file.name.endswith('.xlsx'):
            df1 = pd.read_excel(uploaded_file)
        
        # Dropdown with options for unique names
        unique_names = ['ALL'] + sorted(df1['Name'].unique())
        selected_name = st.selectbox("Filter by Name", unique_names)
        
        # Filter dataframe based on selected name
        filtered_df1 = df1
        if selected_name != 'ALL':  filtered_df1 = df1[df1['Name'] == selected_name]
        
        # Create a new dataframe with an extra column
        df2 = filtered_df1.copy()
        df2, df3, output_file_stream = run_backend(df2)
        
        # Create tabs
        tab1, tab2 = st.tabs(["OUTPUT", "GIVEN DATA"])
        
        with tab1:
            st.markdown("""<h3 style='text-align: center; color: #FFFFFF;'>Output</h3>""", unsafe_allow_html=True)
            st.dataframe(df2, use_container_width=True, hide_index=True)  # Display output dataframe without index


            # Abnotmal Hours
            st.markdown("""<BR><h3 style='text-align: center; color: #FFFFFF;'>Shifts With Abnormal Number of Hours</h3>""", unsafe_allow_html=True)
            st.dataframe(df3, use_container_width=True, hide_index=True)  # Display output dataframe without index


            # Bar Chart
            st.markdown("""<BR><h3 style='text-align: center; color: #FFFFFF;'>Time Spent Per Shift</h3>""", unsafe_allow_html=True)
            shift_hours = df2.groupby('WORK DAY')['HOURS'].sum().reset_index()
            st.bar_chart(shift_hours.set_index('WORK DAY')['HOURS'], use_container_width=True)


            # Add download button
            st.download_button(
                label="Download Excel Output File",
                data=output_file_stream,
                file_name='Output.xlsx',
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                help="Click to download the Output.xlsx file"
            )

            
        with tab2:
            st.markdown("""<h3 style='text-align: center; color: #FFFFFF;'>Original Data</h3>""", unsafe_allow_html=True)
            st.dataframe(filtered_df1, use_container_width=True, hide_index=True)  # Display filtered dataframe without index