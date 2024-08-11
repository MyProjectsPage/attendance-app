#####################################################################################
# VINAVIL ATTENDANCE APPLICATION
# SHADY555@GMAIL.COM
# AUGUST2024
#####################################################################################


#####################################################################################
# BACKEND CALCULATION
#####################################################################################


#import os
#import subprocess
import pandas as pd
#import win32com.client
pd.set_option('display.max_rows', None, 'display.max_columns', None, 'display.max_colwidth', None)
pd.set_option('display.precision', 2)
pd.options.display.width = 0


def create_xl_file(df, output_file = 'Output.xlsx', open_file = False):

    import pandas as pd
    from openpyxl import Workbook
    from openpyxl import load_workbook
    from openpyxl.styles import Font
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.worksheet.table import Table, TableStyleInfo
    

    df.to_excel(output_file, index=False, sheet_name='Output')
    wb = load_workbook(output_file)
    ws = wb.active

    # Create a table and add it to the worksheet
    tab = Table(displayName="Table1", ref=ws.dimensions)
    style = TableStyleInfo(name="TableStyleLight9", showFirstColumn=False, showLastColumn=False, showRowStripes=False, showColumnStripes=True)
    tab.tableStyleInfo = style
    ws.add_table(tab)
    ws.freeze_panes = 'A2'

    # Set font to Calibri size 9 for all cells
    font = Font(name='Calibri', size=9)
    for row in ws.iter_rows():
        for cell in row: cell.font = font

    # Set header font to white
    white_font = Font(color="FFFFFF")
    for cell in ws[1]:  # First header row
        cell.font = Font(name='Calibri', size=9, color="FFFFFF")

    # Auto-fit column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    # Save the workbook
    wb.save(output_file)

    # Open Excel application
    if open_file:
        script_dir = os.path.dirname(os.path.abspath(__file__))
        output_file_path = script_dir + '\\' + output_file
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = True  # Make Excel visible
        workbook = excel.Workbooks.Open(output_file_path)


def delete_irrelivant_entries(df):

    # Initialize an empty list to store the indices to keep
    indices_to_keep = [0]  # Always keep the first row

    # Loop through the DataFrame starting from the second row
    for i in range(1, len(df)):
        # Compare the current row's 'state' with the previous row's 'state'
        if df.iloc[i]['state'] != df.iloc[i - 1]['state']: indices_to_keep.append(i)

    # Create a new DataFrame with the rows to keep
    df = df.iloc[indices_to_keep].reset_index(drop=True)
    return df

def merge_ins_outs(df):
    df_ins = df.iloc[::2].reset_index(drop=True)  # odd index
    df_outs  = df.iloc[1::2].reset_index(drop=True)  # even index
    if len(df_ins) != len(df_outs): raise
    
    df = df_ins[['id', 'name']]
    df.loc[:, ['state_in', 'date_time_in']] = df_ins[['state', 'date_time']].values
    df.loc[:, ['state_out', 'date_time_out']] = df_outs[['state', 'date_time']].values
    return df

def calc_time_spent(df):
    df = df.copy(deep=True) # To avoid copy warning
    df['time_spent'] = df['date_time_out'] - df['date_time_in']
    df['hours'] = df['time_spent'].dt.total_seconds() / 3600
    df['time_spent'] = df['time_spent'].astype(str)
    df['time_spent'] = df['time_spent'].str[-8:-3]
    df['hours'] = df['hours'].round(1)

    for item in ['date_time_in', 'date_time_out']:
        df[item] = df[item].dt.strftime('%d-%m-%Y %I:%M %p')  # AM PM Format
    df['name'] = df['name'].str.upper()
    df.insert(2, 'shift', range(1, len(df) + 1))
    df.insert(3, 'work day', df['date_time_in'])
    df['work day'] = pd.to_datetime(df['work day'], format='%d-%m-%Y %I:%M %p').dt.strftime('%d-%a').str.upper()
    df.columns = df.columns.str.upper()
    
    return df

    

def run_backend(df):

    #####################################################################################
    # BCKEND PROCESSING - MAIN
    #####################################################################################

    dfr = pd.DataFrame()
    #df = pd.read_excel('attendance.xlsx')
    df.columns = 'id name date_time state'.split()
    #df = df[df['name'] == 'Adnan Mohamed Abdul Hami']

    df['date_time'] = pd.to_datetime(df['date_time'], format='%m/%d/%Y %I:%M %p')  
    df = df.sort_values(by=df.columns.tolist())
    df = df[['id', 'name', 'state', 'date_time']]
    df_all = df.copy(deep=True)

    names = (df['name'].unique())

    for name in names:
        df = df_all[df_all['name'] == name]
        while df.iloc[0]['state']  != 'C/In': df = df.iloc[1:].reset_index(drop=True)
        while df.iloc[-1]['state'] != 'C/Out': df = df.iloc[:-1].reset_index(drop=True)
        df = delete_irrelivant_entries(df)
        df = merge_ins_outs(df)
        df = calc_time_spent(df)
        dfr = pd.concat([dfr, df], ignore_index=True)

    create_xl_file(dfr)
    print('\n'*3)
    print(dfr)

    # Abnormal hours (e.g. someone working 23 hours)
    df_abnormal = dfr[~dfr['HOURS'].between(2, 11)]

    return dfr, df_abnormal










#####################################################################################
# FRONTEND STREAMLIT APP
#####################################################################################


import streamlit as st
import pandas as pd

# Set page configuration
st.set_page_config(page_title='Attendance Calculator', page_icon='📊', layout='wide', initial_sidebar_state='expanded')

# Use dark theme and set background image from URL
st.markdown(
    """
    <style>
    .reportview-container {
        background: none;
    }
    .main {
        background: url('https://wallpapercave.com/wp/wp4312403.jpg') no-repeat center center fixed;
        background: url('https://img.freepik.com/premium-photo/abstract-wave-element-design-blue-curve-light-lines-background-digital-frequency-track-equalizer-generative-ai_1423-11938.jpg?w=826') no-repeat center center fixed;
        background: url('https://wallpapercave.com/wp/wp4312424.jpg') no-repeat center center fixed;
        background-size: cover;
    }
    </style>
    """,
    unsafe_allow_html=True
)

# Title
st.title('Attendance Calculator')

# File upload
uploaded_file = st.file_uploader("Upload your file", type=["csv", "xlsx"])

if uploaded_file is not None:
    # Read file into dataframe
    if uploaded_file.name.endswith('.csv'):
        df1 = pd.read_csv(uploaded_file)
    elif uploaded_file.name.endswith('.xlsx'):
        df1 = pd.read_excel(uploaded_file)
    
    # Dropdown with options for unique names
    unique_names = ['ALL'] + sorted(df1['Name'].unique())
    selected_name = st.selectbox("Filter by Name", unique_names)
    
    # Filter dataframe based on selected name
    filtered_df1 = df1
    if selected_name != 'ALL':  filtered_df1 = df1[df1['Name'] == selected_name]
    
    # Create a new dataframe with an extra column
    df2 = filtered_df1.copy()
    df2, df3 = run_backend(df2)
    
    # Create tabs
    tab1, tab2 = st.tabs(["INPUT", "OUTPUT"])
    
    with tab1:
        st.header("Original Data")
        st.dataframe(filtered_df1, use_container_width=True, hide_index=True)  # Display filtered dataframe without index
    
    with tab2:
        st.header("Output")
        st.dataframe(df2, use_container_width=True, hide_index=True)  # Display output dataframe without index


        # Title for the bar chart
        st.markdown("""<BR><h3 style='text-align: center; color: #FFFFFF;'>Shifts With Abnormal Number of Hours</h3>""", unsafe_allow_html=True)
        st.dataframe(df3, use_container_width=True, hide_index=True)  # Display output dataframe without index


        # Title for the bar chart
        st.markdown("""<BR><h3 style='text-align: center; color: #FFFFFF;'>Time Spent Per Shift</h3>""", unsafe_allow_html=True)
        
        # Bar chart for time spent per shift
        shift_hours = df2.groupby('WORK DAY')['HOURS'].sum().reset_index()
        st.bar_chart(shift_hours.set_index('WORK DAY')['HOURS'], use_container_width=True)
